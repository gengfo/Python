import openpyxl

wb = openpyxl.Workbook()


sheet1 = wb.create_sheet(index=0, title='first sheet')
sheet1['A1'] = 'A1 in first sheet'
sheet1.cell(row=1, column=1).value = 'updated'

sheet2 = wb.create_sheet(index=2, title='middle sheet')
sheet2['A1'] = 'A1 in middle sheet'

wb.save("excel6.xlsx")



