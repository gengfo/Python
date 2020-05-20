import openpyxl

wb = openpyxl.load_workbook('test.xlsx')

print(type(wb))

sheetnames = wb.get_sheet_names()

print('sheetnames: {}'.format(sheetnames))

sheet = wb.get_sheet_by_name('Sheet1')

c1 = sheet.cell(row=1,column=1)

print ('type of c1')
print(type (c1))
print(c1.value)

c = sheet['A1']

print ('type of c')
print (type(c))

print ('row of c')
print (str(c.row))

print('A1')
print  ( sheet['B1'].value)

print(sheetnames)